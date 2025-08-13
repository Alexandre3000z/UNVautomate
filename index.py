import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from config.driverConfig import get_webdriver
from services.login import loginExecute
from services.setup import horarioConfig, OSDconfig, audioVideoConfig, redeConfig, serviceConfig, WDRconfig
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

def main():
    # Ler o arquivo Excel
    df = pd.read_excel('FonteDados.xlsx')
    
    # Adicionar coluna 'Concluida' se não existir
    if 'CONCLUIDA' not in df.columns:
        df['CONCLUIDA'] = ""

    # Definir colunas
    colunas = ['IP ANTIGO', 'IP NOVO', 'MASCARA', 'GATEWAY', 'NOME']
    
    # Inicializar o driver
    driver = get_webdriver()
    if not driver:
        print("Erro ao iniciar o driver!")
        return

    # Iterar sobre as linhas
    for index, row in df[colunas].iterrows():
        ipAntigo, ipNovo, mascara, gateway, nomeCamera = row['IP ANTIGO'], row['IP NOVO'], row['MASCARA'], row['GATEWAY'], row['NOME']
        
        try:
            # Acessar o IP antigo
            driver.get(f"http://{ipAntigo}")
            time.sleep(2)

            # Esperar e mudar para o iframe
            WebDriverWait(driver, 10).until(
                EC.frame_to_be_available_and_switch_to_it((By.XPATH, "//iframe[@id='banner']"))
            )

            # Executar as configurações
            loginExecute(driver)

            setupButton = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="nav_config"]'))
            )
            setupButton.click()

            horarioConfig(driver)
            OSDconfig(driver, nomeCamera)
            serviceConfig(driver)
            audioVideoConfig(driver)
            WDRconfig(driver)
            redeConfig(driver, ipNovo, mascara, gateway)
            
            # Se chegou até aqui sem erro, marca como "Sim"
            df.at[index, 'CONCLUIDA'] = "Sim"
            
        except Exception as e:
            print(f"Erro na linha {index}: {e}")
            df.at[index, 'CONCLUIDA'] = "Não"
        
        # Voltar ao contexto principal (fora do iframe)
        driver.switch_to.default_content()

    # Salvar o DataFrame no Excel
    df.to_excel('FonteDados.xlsx', index=False)

    # Aplicar formatação com openpyxl
    wb = load_workbook('FonteDados.xlsx')
    ws = wb.active
    col_idx = df.columns.get_loc('CONCLUIDA') + 1  # Índice da coluna Concluida
    
    verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    for row_idx, valor in enumerate(df['CONCLUIDA'], start=2):  # Começa na linha 2 (após cabeçalho)
        celula = ws.cell(row=row_idx, column=col_idx)
        celula.fill = verde if valor == "Sim" else vermelho

    # Salvar o arquivo com formatação
    wb.save('FonteDados.xlsx')

    # Fechar o driver
    driver.quit()

if __name__ == "__main__":
    main()